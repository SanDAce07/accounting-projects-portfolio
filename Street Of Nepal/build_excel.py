"""
Builds the Streets of Nepal Variance Analysis Excel workbook.
All variance figures are computed via live Excel formulas (not hardcoded),
per the standard cost cards and daily actuals.
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, Reference

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "outputs"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, color="0000FF", size=10)          # hardcoded inputs
BLACK = Font(name=FONT_NAME, color="000000", size=10)          # formulas
BOLD_BLACK = Font(name=FONT_NAME, color="000000", size=10, bold=True)
WHITE_BOLD = Font(name=FONT_NAME, color="FFFFFF", size=11, bold=True)
TITLE_FONT = Font(name=FONT_NAME, color="000000", size=14, bold=True)
SUBTITLE_FONT = Font(name=FONT_NAME, color="595959", size=10, italic=True)

HEADER_FILL = PatternFill("solid", fgColor="2E5395")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CURRENCY = '$#,##0.00;($#,##0.00);"-"'
CURRENCY0 = '$#,##0;($#,##0);"-"'
NUM3 = '0.000;(0.000);"-"'
PCT1 = '0.0%'

def style_header_row(ws, row, n_cols, start_col=1):
    for c in range(start_col, start_col + n_cols):
        cell = ws.cell(row=row, column=c)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

wb = Workbook()

# =================================================================
# SHEET 1: COVER / README
# =================================================================
ws0 = wb.active
ws0.title = "Cover"
ws0.sheet_view.showGridLines = False
ws0["B2"] = "Streets of Nepal"
ws0["B2"].font = Font(name=FONT_NAME, size=22, bold=True, color="2E5395")
ws0["B3"] = "Standard Costing Variance Analysis — April 2026"
ws0["B3"].font = Font(name=FONT_NAME, size=13, bold=True)
ws0["B5"] = "Prepared by: Sandesh Lama Tamang"
ws0["B5"].font = BLACK
ws0["B6"] = "Purpose: Analyze material and labor variances against standard cost cards for four signature dishes, and surface patterns that may indicate operational or control issues."
ws0["B6"].font = BLACK
ws0["B6"].alignment = Alignment(wrap_text=True)
ws0.merge_cells("B6:H6")
ws0.row_dimensions[6].height = 30

notes = [
    ("Sheet", "Contents"),
    ("Standard Cost Cards", "Standard material quantity/price and labor minutes/rate per dish"),
    ("Daily Data", "Raw daily order volume and actual material/labor figures"),
    ("Variance Calcs", "Formula-driven MPV, MQV, LRV, LEV for every dish-day"),
    ("Dashboard", "Monthly summary by dish, conditional formatting, chart, and red-flag callouts"),
]
ws0["B9"] = "Workbook Contents"
ws0["B9"].font = BOLD_BLACK
for i, (a, b) in enumerate(notes):
    r = 10 + i
    ws0.cell(row=r, column=2, value=a).font = BOLD_BLACK if i == 0 else BLACK
    ws0.cell(row=r, column=3, value=b).font = BOLD_BLACK if i == 0 else BLACK
    ws0.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)

ws0["B17"] = "Color key: Blue = input cell  |  Black = formula  |  Yellow = key assumption  |  Red fill = unfavorable / flagged"
ws0["B17"].font = SUBTITLE_FONT
ws0.merge_cells("B17:H17")
autosize(ws0, {"A": 3, "B": 20, "C": 60})

# =================================================================
# SHEET 2: STANDARD COST CARDS
# =================================================================
ws1 = wb.create_sheet("Standard Cost Cards")
ws1.sheet_view.showGridLines = False
ws1["B2"] = "Standard Cost Cards"
ws1["B2"].font = TITLE_FONT
ws1["B3"] = "All blue cells are hardcoded standard-cost assumptions, set by management at the start of the period."
ws1["B3"].font = SUBTITLE_FONT
ws1.merge_cells("B3:H3")

headers = ["Dish", "Std Material Qty (lb/order)", "Std Material Price ($/lb)",
           "Std Material Cost/Order", "Std Labor Minutes/Order", "Std Labor Rate ($/hr)",
           "Std Labor Cost/Order", "Std Total Cost/Order"]
hdr_row = 5
for i, h in enumerate(headers):
    ws1.cell(row=hdr_row, column=2 + i, value=h)
style_header_row(ws1, hdr_row, len(headers), start_col=2)

dishes_data = [
    ("Chicken Momos", 0.35, 3.80, 6.0, 13.50),
    ("Veg Thukpa", 0.60, 2.10, 8.0, 13.50),
    ("Chow Mein (Chicken)", 0.50, 3.20, 7.0, 13.50),
    ("Lamb Sekuwa", 0.45, 6.50, 10.0, 14.50),
]

for i, (name, sq, sp, sm, sr) in enumerate(dishes_data):
    r = hdr_row + 1 + i
    ws1.cell(row=r, column=2, value=name).font = BLUE
    ws1.cell(row=r, column=3, value=sq).font = BLUE
    ws1.cell(row=r, column=3).number_format = NUM3
    ws1.cell(row=r, column=4, value=sp).font = BLUE
    ws1.cell(row=r, column=4).number_format = CURRENCY
    ws1.cell(row=r, column=5, value=f"=C{r}*D{r}").font = BLACK
    ws1.cell(row=r, column=5).number_format = CURRENCY
    ws1.cell(row=r, column=6, value=sm).font = BLUE
    ws1.cell(row=r, column=6).number_format = NUM3
    ws1.cell(row=r, column=7, value=sr).font = BLUE
    ws1.cell(row=r, column=7).number_format = CURRENCY
    ws1.cell(row=r, column=8, value=f"=(F{r}/60)*G{r}").font = BLACK
    ws1.cell(row=r, column=8).number_format = CURRENCY
    ws1.cell(row=r, column=9, value=f"=E{r}+H{r}").font = BOLD_BLACK
    ws1.cell(row=r, column=9).number_format = CURRENCY
    for c in range(2, 10):
        ws1.cell(row=r, column=c).border = BORDER

autosize(ws1, {"A": 2, "B": 22, "C": 16, "D": 16, "E": 16, "F": 16, "G": 14, "H": 16, "I": 16})

# =================================================================
# SHEET 3: DAILY DATA (raw actuals)
# =================================================================
df = pd.read_csv(DATA_DIR / "daily_actuals.csv")
ws2 = wb.create_sheet("Daily Data")
ws2.sheet_view.showGridLines = False
ws2["A1"] = "Daily Actuals — April 2026 (26 operating days x 4 dishes)"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = "Source: POS order counts and kitchen/purchasing logs (blue = actual recorded inputs)."
ws2["A2"].font = SUBTITLE_FONT

cols = ["Dish", "Day", "Orders", "Std_Material_Qty_per_Order", "Std_Material_Price",
        "Actual_Material_Price", "Actual_Material_Qty_per_Order", "Std_Labor_Minutes_per_Order",
        "Std_Labor_Rate", "Actual_Labor_Minutes_per_Order", "Actual_Labor_Rate"]
display_headers = ["Dish", "Day", "Orders", "Std Mat Qty/Order", "Std Mat Price",
                    "Actual Mat Price", "Actual Mat Qty/Order", "Std Labor Min/Order",
                    "Std Labor Rate", "Actual Labor Min/Order", "Actual Labor Rate"]

hdr_row = 4
for i, h in enumerate(display_headers):
    ws2.cell(row=hdr_row, column=1 + i, value=h)
style_header_row(ws2, hdr_row, len(display_headers), start_col=1)

for i, row in df.iterrows():
    r = hdr_row + 1 + i
    ws2.cell(row=r, column=1, value=row["Dish"]).font = BLUE
    ws2.cell(row=r, column=2, value=int(row["Day"])).font = BLUE
    ws2.cell(row=r, column=3, value=int(row["Orders"])).font = BLUE
    ws2.cell(row=r, column=4, value=row["Std_Material_Qty_per_Order"]).font = BLUE
    ws2.cell(row=r, column=4).number_format = NUM3
    ws2.cell(row=r, column=5, value=row["Std_Material_Price"]).font = BLUE
    ws2.cell(row=r, column=5).number_format = CURRENCY
    ws2.cell(row=r, column=6, value=row["Actual_Material_Price"]).font = BLUE
    ws2.cell(row=r, column=6).number_format = CURRENCY
    ws2.cell(row=r, column=7, value=row["Actual_Material_Qty_per_Order"]).font = BLUE
    ws2.cell(row=r, column=7).number_format = NUM3
    ws2.cell(row=r, column=8, value=row["Std_Labor_Minutes_per_Order"]).font = BLUE
    ws2.cell(row=r, column=8).number_format = NUM3
    ws2.cell(row=r, column=9, value=row["Std_Labor_Rate"]).font = BLUE
    ws2.cell(row=r, column=9).number_format = CURRENCY
    ws2.cell(row=r, column=10, value=row["Actual_Labor_Minutes_per_Order"]).font = BLUE
    ws2.cell(row=r, column=10).number_format = NUM3
    ws2.cell(row=r, column=11, value=row["Actual_Labor_Rate"]).font = BLUE
    ws2.cell(row=r, column=11).number_format = CURRENCY

n_data_rows = len(df)
last_data_row = hdr_row + n_data_rows
autosize(ws2, {get_column_letter(i): 16 for i in range(1, 12)})
ws2.column_dimensions["A"].width = 22
ws2.freeze_panes = "A5"

# =================================================================
# SHEET 4: VARIANCE CALCS (live formulas referencing Daily Data)
# =================================================================
ws3 = wb.create_sheet("Variance Calcs")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "Variance Calculations (all formulas; negative = favorable)"
ws3["A1"].font = TITLE_FONT
ws3["A2"] = "MPV = (Actual Price - Std Price) x Actual Qty   |   MQV = (Actual Qty - Std Qty) x Std Price"
ws3["A2"].font = SUBTITLE_FONT
ws3["A3"] = "LRV = (Actual Rate - Std Rate) x Actual Hours    |   LEV = (Actual Hours - Std Hours) x Std Rate"
ws3["A3"].font = SUBTITLE_FONT

vc_headers = ["Dish", "Day", "Orders", "Actual Mat Qty Total", "Std Mat Qty Total",
              "Actual Labor Hrs Total", "Std Labor Hrs Total",
              "MPV", "MQV", "LRV", "LEV", "Total Material Var", "Total Labor Var", "Grand Total Var"]
hdr_row3 = 5
for i, h in enumerate(vc_headers):
    ws3.cell(row=hdr_row3, column=1 + i, value=h)
style_header_row(ws3, hdr_row3, len(vc_headers), start_col=1)

for i in range(n_data_rows):
    r = hdr_row3 + 1 + i
    dd_r = hdr_row + 1 + i  # corresponding row in Daily Data
    ws3.cell(row=r, column=1, value=f"='Daily Data'!A{dd_r}").font = BLACK
    ws3.cell(row=r, column=2, value=f"='Daily Data'!B{dd_r}").font = BLACK
    ws3.cell(row=r, column=3, value=f"='Daily Data'!C{dd_r}").font = BLACK
    # Actual material qty total = actual qty/order * orders
    ws3.cell(row=r, column=4, value=f"='Daily Data'!G{dd_r}*'Daily Data'!C{dd_r}").font = BLACK
    ws3.cell(row=r, column=4).number_format = NUM3
    # Std material qty total = std qty/order * orders
    ws3.cell(row=r, column=5, value=f"='Daily Data'!D{dd_r}*'Daily Data'!C{dd_r}").font = BLACK
    ws3.cell(row=r, column=5).number_format = NUM3
    # Actual labor hours total = (actual min/order * orders) / 60
    ws3.cell(row=r, column=6, value=f"=('Daily Data'!J{dd_r}*'Daily Data'!C{dd_r})/60").font = BLACK
    ws3.cell(row=r, column=6).number_format = NUM3
    # Std labor hours total = (std min/order * orders) / 60
    ws3.cell(row=r, column=7, value=f"=('Daily Data'!H{dd_r}*'Daily Data'!C{dd_r})/60").font = BLACK
    ws3.cell(row=r, column=7).number_format = NUM3
    # MPV = (Actual Price - Std Price) * Actual Qty Total
    ws3.cell(row=r, column=8, value=f"=('Daily Data'!F{dd_r}-'Daily Data'!E{dd_r})*D{r}").font = BLACK
    ws3.cell(row=r, column=8).number_format = CURRENCY
    # MQV = (Actual Qty Total - Std Qty Total) * Std Price
    ws3.cell(row=r, column=9, value=f"=(D{r}-E{r})*'Daily Data'!E{dd_r}").font = BLACK
    ws3.cell(row=r, column=9).number_format = CURRENCY
    # LRV = (Actual Rate - Std Rate) * Actual Hours Total
    ws3.cell(row=r, column=10, value=f"=('Daily Data'!K{dd_r}-'Daily Data'!I{dd_r})*F{r}").font = BLACK
    ws3.cell(row=r, column=10).number_format = CURRENCY
    # LEV = (Actual Hours Total - Std Hours Total) * Std Rate
    ws3.cell(row=r, column=11, value=f"=(F{r}-G{r})*'Daily Data'!I{dd_r}").font = BLACK
    ws3.cell(row=r, column=11).number_format = CURRENCY
    ws3.cell(row=r, column=12, value=f"=H{r}+I{r}").font = BOLD_BLACK
    ws3.cell(row=r, column=12).number_format = CURRENCY
    ws3.cell(row=r, column=13, value=f"=J{r}+K{r}").font = BOLD_BLACK
    ws3.cell(row=r, column=13).number_format = CURRENCY
    ws3.cell(row=r, column=14, value=f"=L{r}+M{r}").font = BOLD_BLACK
    ws3.cell(row=r, column=14).number_format = CURRENCY

last_vc_row = hdr_row3 + n_data_rows
# Conditional formatting: red for unfavorable (positive), green for favorable (negative)
for col_letter in ["H", "I", "J", "K", "L", "M", "N"]:
    rng = f"{col_letter}{hdr_row3+1}:{col_letter}{last_vc_row}"
    ws3.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=RED_FILL))
    ws3.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=GREEN_FILL))

autosize(ws3, {get_column_letter(i): 15 for i in range(1, 15)})
ws3.column_dimensions["A"].width = 22
ws3.freeze_panes = "A6"

# =================================================================
# SHEET 5: DASHBOARD (monthly summary by dish + flags)
# =================================================================
ws4 = wb.create_sheet("Dashboard")
ws4.sheet_view.showGridLines = False
ws4["B2"] = "Monthly Variance Dashboard — April 2026"
ws4["B2"].font = TITLE_FONT
ws4["B3"] = "Negative (green) = favorable to standard. Positive (red) = unfavorable to standard."
ws4["B3"].font = SUBTITLE_FONT

dash_headers = ["Dish", "Total Orders", "MPV", "MQV", "LRV", "LEV", "Total Material Var", "Total Labor Var", "Grand Total Var"]
hdr_row4 = 5
for i, h in enumerate(dash_headers):
    ws4.cell(row=hdr_row4, column=2 + i, value=h)
style_header_row(ws4, hdr_row4, len(dash_headers), start_col=2)

dish_names = [d[0] for d in dishes_data]
for i, dish in enumerate(dish_names):
    r = hdr_row4 + 1 + i
    ws4.cell(row=r, column=2, value=dish).font = BOLD_BLACK
    # SUMIF against Variance Calcs sheet
    ws4.cell(row=r, column=3, value=f"=SUMIF('Variance Calcs'!$A${hdr_row3+1}:$A${last_vc_row},B{r},'Variance Calcs'!$H${hdr_row3+1}:$H${last_vc_row})").font = BLACK
    ws4.cell(row=r, column=4, value=f"=SUMIF('Variance Calcs'!$A${hdr_row3+1}:$A${last_vc_row},B{r},'Variance Calcs'!$I${hdr_row3+1}:$I${last_vc_row})").font = BLACK
    ws4.cell(row=r, column=5, value=f"=SUMIF('Variance Calcs'!$A${hdr_row3+1}:$A${last_vc_row},B{r},'Variance Calcs'!$J${hdr_row3+1}:$J${last_vc_row})").font = BLACK
    ws4.cell(row=r, column=6, value=f"=SUMIF('Variance Calcs'!$A${hdr_row3+1}:$A${last_vc_row},B{r},'Variance Calcs'!$K${hdr_row3+1}:$K${last_vc_row})").font = BLACK
    ws4.cell(row=r, column=7, value=f"=SUMIF('Daily Data'!$A${hdr_row+1}:$A${last_data_row},B{r},'Daily Data'!$C${hdr_row+1}:$C${last_data_row})").font = BLACK
    # reposition: column 2 already orders -- fix mapping below
    ws4.cell(row=r, column=8, value=f"=C{r}+D{r}").font = BOLD_BLACK
    ws4.cell(row=r, column=9, value=f"=E{r}+F{r}").font = BOLD_BLACK
    ws4.cell(row=r, column=10, value=f"=H{r}+I{r}").font = BOLD_BLACK
    for c in range(3, 11):
        ws4.cell(row=r, column=c).number_format = CURRENCY
    ws4.cell(row=r, column=7).number_format = '#,##0'

# Fix: column G should be Total Orders (count), not currency; relabel correctly
# Re-do headers to match actual column layout used above
ws4.cell(row=hdr_row4, column=2, value="Dish")
ws4.cell(row=hdr_row4, column=3, value="MPV")
ws4.cell(row=hdr_row4, column=4, value="MQV")
ws4.cell(row=hdr_row4, column=5, value="LRV")
ws4.cell(row=hdr_row4, column=6, value="LEV")
ws4.cell(row=hdr_row4, column=7, value="Total Orders")
ws4.cell(row=hdr_row4, column=8, value="Total Material Var")
ws4.cell(row=hdr_row4, column=9, value="Total Labor Var")
ws4.cell(row=hdr_row4, column=10, value="Grand Total Var")
style_header_row(ws4, hdr_row4, 9, start_col=2)

last_dash_row = hdr_row4 + len(dish_names)
ws4.cell(row=last_dash_row + 1, column=2, value="Company Total").font = BOLD_BLACK
for c in range(3, 11):
    col_l = get_column_letter(c)
    ws4.cell(row=last_dash_row + 1, column=c, value=f"=SUM({col_l}{hdr_row4+1}:{col_l}{last_dash_row})").font = BOLD_BLACK
    ws4.cell(row=last_dash_row + 1, column=c).number_format = CURRENCY if c != 7 else '#,##0'
    ws4.cell(row=last_dash_row + 1, column=c).border = Border(top=Side(style="double"))

for col_letter in ["C", "D", "E", "F", "H", "I", "J"]:
    rng = f"{col_letter}{hdr_row4+1}:{col_letter}{last_dash_row}"
    ws4.conditional_formatting.add(rng, CellIsRule(operator="greaterThan", formula=["0"], fill=RED_FILL))
    ws4.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0"], fill=GREEN_FILL))

for c in range(2, 11):
    for r in range(hdr_row4, last_dash_row + 2):
        ws4.cell(row=r, column=c).border = BORDER

autosize(ws4, {get_column_letter(i): 16 for i in range(1, 11)})
ws4.column_dimensions["B"].width = 22

# --- Chart: Grand total variance by dish ---
chart = BarChart()
chart.title = "Total Variance by Dish ($)"
chart.legend = None
chart.y_axis.title = "Variance ($)"
chart.x_axis.title = "Dish"
data_ref = Reference(ws4, min_col=10, max_col=10, min_row=hdr_row4, max_row=last_dash_row)
cats_ref = Reference(ws4, min_col=2, min_row=hdr_row4 + 1, max_row=last_dash_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 18
chart.height = 9
ws4.add_chart(chart, "B" + str(last_dash_row + 4))

# --- Red flag summary box ---
flag_row = last_dash_row + 22
ws4.cell(row=flag_row, column=2, value="Red Flags Identified (see Audit Memo for full detail)").font = BOLD_BLACK
ws4.cell(row=flag_row, column=2).fill = SECTION_FILL
ws4.merge_cells(start_row=flag_row, start_column=2, end_row=flag_row, end_column=9)

flags_df = pd.read_csv(DATA_DIR / "red_flags.csv")
fr = flag_row + 1
ws4.cell(row=fr, column=2, value="Risk").font = BOLD_BLACK
ws4.cell(row=fr, column=3, value="Test").font = BOLD_BLACK
ws4.cell(row=fr, column=4, value="Dish").font = BOLD_BLACK
ws4.cell(row=fr, column=5, value="Detail").font = BOLD_BLACK
ws4.merge_cells(start_row=fr, start_column=5, end_row=fr, end_column=9)
for c in range(2, 10):
    ws4.cell(row=fr, column=c).fill = SECTION_FILL
    ws4.cell(row=fr, column=c).border = BORDER

for i, frow in flags_df.iterrows():
    r = fr + 1 + i
    ws4.cell(row=r, column=2, value=frow["risk"]).font = BLACK
    ws4.cell(row=r, column=2).fill = RED_FILL if frow["risk"] == "High" else YELLOW_FILL
    ws4.cell(row=r, column=3, value=frow["test"]).font = BLACK
    ws4.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.cell(row=r, column=4, value=frow["dish"]).font = BLACK
    ws4.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.cell(row=r, column=5, value=frow["detail"]).font = BLACK
    ws4.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)
    for c in range(2, 10):
        ws4.cell(row=r, column=c).border = BORDER
    ws4.row_dimensions[r].height = 45

ws4.column_dimensions["C"].width = 24
ws4.column_dimensions["D"].width = 24

output_path = OUTPUT_DIR / "Streets_of_Nepal_Variance_Analysis.xlsx"
wb.save(output_path)
print(f"Saved {output_path}")
